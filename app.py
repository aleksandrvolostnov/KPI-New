from flask import Flask, render_template, redirect, url_for, request, flash, send_file, jsonify, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, login_user, login_required, logout_user, UserMixin, current_user
from datetime import datetime, timedelta
import csv
import io
from calendar import monthcalendar, month_name
from babel.dates import format_date
import locale
import os
from werkzeug.utils import secure_filename
from openpyxl.chart import BarChart, Reference
from werkzeug.utils import secure_filename
import json
from hashlib import md5

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your_secret_key'
app.config['SQLALCHEMY_DATABASE_URI'] = 'postgresql://postgres:1111@localhost/efficiency_control'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

login_manager = LoginManager(app)
login_manager.login_view = 'login'


# Модели
class Role(db.Model):
    __tablename__ = 'roles'
    id = db.Column(db.Integer, primary_key=True)
    role_name = db.Column(db.String(50), nullable=False)


class User(db.Model, UserMixin):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(150), unique=True, nullable=False)
    password = db.Column(db.String(150), nullable=False)
    role_id = db.Column(db.Integer, db.ForeignKey('roles.id'))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    role = db.relationship('Role', backref='users')
    email = db.Column(db.String(150), unique=True, nullable=True)  # Новое поле для почты
    phone = db.Column(db.String(20), nullable=True)  # Новое поле для телефона
    avatar = db.Column(db.String(255), nullable=True)  # Новое поле для аватарки
<<<<<<< HEAD
    tokens = db.Column(db.Integer, default=0)  # Поле для жетонов
=======
>>>>>>> 4c80bc660c306d2a5b2908cf97d88b29abcddb6f

    sent_messages = db.relationship('Message', foreign_keys='Message.sender_id', backref='sender_rel', lazy='dynamic')
    received_messages = db.relationship('Message', foreign_keys='Message.receiver_id', backref='receiver_rel', lazy='dynamic')

<<<<<<< HEAD
class Project(db.Model):
    __tablename__ = 'projects'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(150), nullable=False)
    description = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    owner_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    owner = db.relationship('User', backref='owned_projects', lazy=True)
    tasks = db.relationship('Task', backref='project', lazy='dynamic')
    members = db.relationship('User', secondary='project_members', backref='projects')
    project_members = db.Table('project_members',
    db.Column('project_id', db.Integer, db.ForeignKey('projects.id'), primary_key=True),
    db.Column('user_id', db.Integer, db.ForeignKey('users.id'), primary_key=True)
                               )

class Reminder(db.Model):
    __tablename__ = 'reminders'
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(150), nullable=False)
    description = db.Column(db.Text, nullable=True)
    reminder_date = db.Column(db.DateTime, nullable=False)
    priority = db.Column(db.String(50), default='Низкий')  # Приоритет
    repeat = db.Column(db.String(50), default='Нет')  # Повторение (Нет, Ежедневно, Еженедельно, Ежемесячно)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)

    user = db.relationship('User', backref='reminders')
=======
>>>>>>> 4c80bc660c306d2a5b2908cf97d88b29abcddb6f


class Message(db.Model):
    __tablename__ = 'messages'
    id = db.Column(db.Integer, primary_key=True)
    sender_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    receiver_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    is_group = db.Column(db.Boolean, default=False)
    content = db.Column(db.Text, nullable=True)
    filename = db.Column(db.String(150), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    is_read = db.Column(db.Boolean, default=False)
    parent_message_id = db.Column(db.Integer, db.ForeignKey('messages.id'), nullable=True)  # Поле для родительского сообщения

    # Взаимосвязи
    sender = db.relationship('User', foreign_keys=[sender_id], backref='sent_messages_backref')
    receiver = db.relationship('User', foreign_keys=[receiver_id], backref='received_messages_backref', lazy='joined')
    parent_message = db.relationship('Message', remote_side=[id], backref='replies', lazy='joined')  # Взаимосвязь для ответов

    # Обновление `updated_at` при каждом изменении
    def update_timestamp(self):
        self.updated_at = datetime.utcnow()



class Task(db.Model):
    __tablename__ = 'tasks'
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(150), nullable=False)
    description = db.Column(db.Text, nullable=False)
    priority = db.Column(db.String(50), nullable=False)
    status = db.Column(db.String(50), nullable=False)
    difficulty = db.Column(db.String(50), nullable=False)
    due_date = db.Column(db.DateTime, nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'))  # Создатель задачи
    assigned_to_id = db.Column(db.Integer, db.ForeignKey('users.id'))  # Ответственный
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
<<<<<<< HEAD
    project_id = db.Column(db.Integer, db.ForeignKey('projects.id'), nullable=False)

    parent_task_id = db.Column(db.Integer, db.ForeignKey('tasks.id'))  # Связь с родительской задачей
    parent_task = db.relationship('Task', remote_side=[id], backref='dependent_tasks')  # Родительская задача
=======
>>>>>>> 4c80bc660c306d2a5b2908cf97d88b29abcddb6f

    user = db.relationship('User', foreign_keys=[user_id], backref='tasks_created')
    assigned_to = db.relationship('User', foreign_keys=[assigned_to_id], backref='tasks_assigned')

    def to_dict(self):
        return {
            'id': self.id,
            'title': self.title,
            'description': self.description,
            'priority': self.priority,
            'status': self.status,
            'due_date': self.due_date.strftime('%Y-%m-%d'),
            'assigned_to': self.assigned_to.username if self.assigned_to else None,
<<<<<<< HEAD
            'project': self.project.name if self.project else None,
            'subtasks': [subtask.to_dict() for subtask in self.subtasks],
            'parent_task': self.parent_task.title if self.parent_task else None,
            'dependent_tasks': [task.title for task in self.dependent_tasks]
=======
            'subtasks': [subtask.to_dict() for subtask in self.subtasks]
>>>>>>> 4c80bc660c306d2a5b2908cf97d88b29abcddb6f
        }


class SubTask(db.Model):
    __tablename__ = 'subtasks'
    id = db.Column(db.Integer, primary_key=True)
    task_id = db.Column(db.Integer, db.ForeignKey('tasks.id'), nullable=False)
    title = db.Column(db.String(150), nullable=False)
    start_date = db.Column(db.DateTime, nullable=False)
    end_date = db.Column(db.DateTime, nullable=False)

    task = db.relationship('Task', backref=db.backref('subtasks', lazy=True))

    def to_dict(self):
        return {
            'id': self.id,
            'task_id': self.task_id,
            'title': self.title,
            'start_date': self.start_date.strftime('%Y-%m-%d') if self.start_date else "",
            'end_date': self.end_date.strftime('%Y-%m-%d') if self.end_date else "",
        }


class TaskComments(db.Model):
    __tablename__ = 'task_comments'
    id = db.Column(db.Integer, primary_key=True)
    task_id = db.Column(db.Integer, db.ForeignKey('tasks.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    comment = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    task = db.relationship('Task', backref='task_comments')
    user = db.relationship('User', backref='user_comments')
    def to_dict(self):
        return {
            'id': self.id,
            'task_id': self.task_id,
            'user_id': self.user_id,
            'comment': self.comment,
            "created_at": self.created_at.strftime('%Y-%m-%d %H:%M:%S')
        }
class File(db.Model):
    __tablename__ = 'files'
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(255), nullable=False)
    task_id = db.Column(db.Integer, db.ForeignKey('tasks.id'))  # Внешний ключ для привязки к задаче


@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    if request.method == 'POST':
        # Проверка и обновление почты
        if 'email' in request.form:
            email = request.form['email']
            if email:  # Если email не пустой, обновляем
                current_user.email = email
                flash('Почта обновлена!', 'success')
            else:  # Если email пустой, удаляем его
                current_user.email = None
                flash('Почта удалена!', 'success')

        # Проверка и обновление телефона
        if 'phone' in request.form:
            phone = request.form['phone']
            if phone:  # Если телефон не пустой, обновляем
                current_user.phone = phone
                flash('Телефон обновлен!', 'success')
            else:  # Если телефон пустой, удаляем его
                current_user.phone = None
                flash('Телефон удален!', 'success')

        # Загрузка аватарки
        if 'avatar' in request.files:
            avatar_file = request.files['avatar']
            if avatar_file and allowed_file(avatar_file.filename):  # Проверка на допустимый файл
                avatar_filename = secure_filename(avatar_file.filename)  # Безопасное имя файла
                avatar_path = os.path.join(app.config['UPLOAD_FOLDER'], avatar_filename)

                # Сохранение аватарки в папке uploads
                avatar_file.save(avatar_path)

                # Сохранение имени аватарки в базе данных
                current_user.avatar = avatar_filename
                flash('Аватарка обновлена!', 'success')

        db.session.commit()  # Сохраняем изменения в базе данных
        return redirect(url_for('profile'))  # Перенаправляем на страницу профиля

    return render_template('profile.html', user=current_user)
@app.route('/user/<int:user_id>')
@login_required
def view_user(user_id):
    user = User.query.get_or_404(user_id)
    return render_template('user_profile.html', user=user)

# Логин менеджер
@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


# Главная страница
@app.route('/')
def index():
    return render_template('welcome.html')

# Функция для обработки прочитанных сообщений
def mark_messages_as_read(sender_id, receiver_id):
    unread_messages = Message.query.filter_by(receiver_id=receiver_id, sender_id=sender_id, is_read=False).all()
    for msg in unread_messages:
        msg.is_read = True
        db.session.commit()

# Функция для получения последнего сообщения
def get_last_message(user_id_1, user_id_2):
    last_message = Message.query.filter(
        ((Message.sender_id == user_id_1) & (Message.receiver_id == user_id_2)) |
        ((Message.sender_id == user_id_2) & (Message.receiver_id == user_id_1))
    ).order_by(Message.created_at.desc()).first()

    if last_message:
        return last_message
    return None

# Основной маршрут для чата
@app.route('/chat/<user_id>', methods=['GET', 'POST'])
@login_required
def chat(user_id):
    users = User.query.filter(User.id != current_user.id).all()
    last_messages = {}
    user_has_new_message = {}

    for user in users:
        # Получаем последнее сообщение
        last_message = get_last_message(current_user.id, user.id)

        if last_message:
            last_messages[user.id] = last_message.content if last_message.content else f'Файл: {last_message.filename}'
        else:
            last_messages[user.id] = 'Нет сообщений'

        # Проверка на наличие новых сообщений
        has_new_message = Message.query.filter_by(sender_id=user.id, receiver_id=current_user.id).filter(
            Message.is_read == False
        ).count() > 0

        user_has_new_message[user.id] = has_new_message

    # Сортировка пользователей по времени последнего сообщения (если есть)
    users_sorted = sorted(users, key=lambda user: (
        get_last_message(current_user.id, user.id).created_at if get_last_message(current_user.id, user.id) else datetime.min
    ), reverse=True)

    # Обработка отправки нового сообщения
    if request.method == 'POST':
        content = request.form.get('content')
        file = request.files.get('file')
        parent_message_id = request.form.get('parent_message_id')

        if not content and not file:
            return jsonify({"status": "error", "message": "Сообщение не может быть пустым"})

        filename = None
        if file:
            filename = file.filename
            file.save(f'uploads/{filename}')

        new_message = Message(sender_id=current_user.id, receiver_id=user_id, content=content, filename=filename)

        if parent_message_id:
            new_message.parent_message_id = parent_message_id

        db.session.add(new_message)
        db.session.commit()

        return jsonify({"status": "success", "message": "Сообщение отправлено"})

    # Вот сюда вставляем код для загрузки сообщений
    if user_id == 'group':
        messages = Message.query.filter_by(is_group=True).order_by(Message.created_at.asc()).all()
    else:
        messages = Message.query.filter(
            ((Message.sender_id == current_user.id) & (Message.receiver_id == user_id)) |
            ((Message.sender_id == user_id) & (Message.receiver_id == current_user.id))
        ).order_by(Message.created_at.asc()).options(db.joinedload(Message.parent_message)).all()  # Загрузить родительские сообщения

    # Отмечаем все непрочитанные сообщения как прочитанные
    unread_messages = Message.query.filter_by(receiver_id=current_user.id, sender_id=user_id, is_read=False).all()
    for msg in unread_messages:
        msg.is_read = True
    db.session.commit()

    chat_with = "Общая группа" if user_id == 'group' else User.query.get(user_id).username

    return render_template('chat.html', messages=messages, users=users_sorted, chat_with=chat_with,
                           last_messages=last_messages, user_has_new_message=user_has_new_message)

# Маршрут для отправки нового сообщения
@app.route('/chat/new_message', methods=['POST'])
@login_required
def new_message():
    content = request.form.get('content')
    receiver_id = request.form.get('receiver_id')
    parent_message_id = request.form.get('parent_message_id')  # Получаем ID родительского сообщения

    if content and receiver_id:
        new_message = Message(sender_id=current_user.id, receiver_id=receiver_id, content=content)

        if parent_message_id:
            new_message.parent_message_id = parent_message_id  # Связываем с родительским сообщением

        db.session.add(new_message)
        db.session.commit()

        return jsonify({
            'status': 'success',
            'message': 'Сообщение отправлено',
            'new_message': {
                'sender': current_user.username,
                'content': content
            }
        })

    return jsonify({'status': 'error', 'message': 'Невозможно отправить сообщение'})

# Маршрут для пересылки сообщения
@app.route('/forward', methods=['POST'])
@login_required
def forward_message():
    message_id = request.form.get('message_id')  # Получаем ID сообщения
    recipient_id = request.form.get('recipient_id')  # Получаем ID получателя

    if not message_id or not recipient_id:
        return jsonify({'status': 'error', 'message': 'Неверные данные для пересылки'}), 400

    # Получаем оригинальное сообщение
    message = Message.query.get(message_id)
    if not message:
        return jsonify({'status': 'error', 'message': 'Сообщение не найдено'}), 404

    # Создаем новое сообщение с тем же контентом и файлом для получателя
    forwarded_message = Message(
        sender_id=current_user.id,  # Отправитель — текущий пользователь
        receiver_id=recipient_id,  # Получатель — выбранный пользователь
        content=f"Пересланное сообщение: {message.content}",
        filename=message.filename  # Пересылаем файл, если он есть
    )

    db.session.add(forwarded_message)
    db.session.commit()

    return jsonify({'status': 'success', 'message': 'Сообщение переслано'})


# Маршрут для ответа на сообщение
@app.route('/reply/<message_id>', methods=['POST'])
@login_required
def reply_to_message(message_id):
    content = request.form.get('reply_content')
    original_message = Message.query.get(message_id)

    if original_message and content:
        reply_message = Message(
            sender_id=current_user.id,
            receiver_id=original_message.sender_id,
            content=content,
            parent_message_id=original_message.id  # Связываем с оригинальным сообщением
        )
        db.session.add(reply_message)
        db.session.commit()

        return jsonify({
            'status': 'success',
            'message': 'Ответ отправлен',
            'new_message': {
                'sender': current_user.username,
                'content': content
            }
        })

    return jsonify({'status': 'error', 'message': 'Невозможно отправить ответ'})


@app.route('/uploads/<filename>')
@login_required
def download_file(filename):
    uploads_dir = 'uploads'
    file_path = os.path.join(uploads_dir, filename)
    if not os.path.exists(file_path):
        return jsonify({"status": "error", "message": "Файл не найден"}), 404

    return send_from_directory(uploads_dir, filename)

# Регистрация
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        confirm_password = request.form['confirm_password']
        role_id = request.form['role_id']

        if password != confirm_password:
            flash('Пароли не совпадают!', 'error')
            return redirect(url_for('register'))

        if User.query.filter_by(username=username).first():
            flash('Пользователь с таким именем уже существует!', 'error')
            return redirect(url_for('register'))

        new_user = User(username=username, password=password, role_id=role_id)
        db.session.add(new_user)
        db.session.commit()

        flash('Регистрация прошла успешно!', 'success')
        return redirect(url_for('login'))

    return render_template('register.html')


# Логин
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username, password=password).first()
        if user:
            login_user(user)
            return redirect(url_for('dashboard'))
        else:
            flash('Неверное имя пользователя или пароль!')
    return render_template('login.html')


# Выход из системы
@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('index'))

# Дашборд с задачами
@app.route('/dashboard')
@login_required
def dashboard():
    if current_user.role.role_name == 'Admin':
        tasks = Task.query.all()
    else:
        tasks = Task.query.filter((Task.user_id == current_user.id) | (Task.assigned_to_id == current_user.id)).all()

    # Сортируем задачи по сроку дедлайна (ближайшие сначала)
    tasks_sorted = sorted(tasks, key=lambda x: x.due_date)

    # Берем только первые 3 задачи с ближайшими дедлайнами
    upcoming_tasks = tasks_sorted[:3]

    # Аналитика
    total_tasks = len(tasks)
    low_priority_count = Task.query.filter_by(priority='Низкий').count()
    medium_priority_count = Task.query.filter_by(priority='Средний').count()
    high_priority_count = Task.query.filter_by(priority='Высокий').count()

    # Количество задач по сложности (исходя из числовых значений)
    easy_count = Task.query.filter_by(difficulty=1).count()
    medium_count = Task.query.filter_by(difficulty=2).count()
    hard_count = Task.query.filter_by(difficulty=3).count()

    # Для вычисления средней сложности задач
    average_difficulty = db.session.query(db.func.avg(Task.difficulty)).scalar()

    return render_template('dashboard.html',
                           tasks=upcoming_tasks,  # передаем только 3 задачи
                           total_tasks=total_tasks,
                           low_priority_count=low_priority_count,
                           medium_priority_count=medium_priority_count,
                           high_priority_count=high_priority_count,
                           average_difficulty=average_difficulty,
                           easy_count=easy_count,
                           medium_count=medium_count,
                           hard_count=hard_count)



# Путь для сохранения загружаемых файлов
UPLOAD_FOLDER = os.path.join(os.getcwd(), 'uploads')
ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'gif', 'docx', 'xlsx', 'xls', 'doc', 'csv'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Функция проверки допустимого расширения файла
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/task/new', methods=['GET', 'POST'])
@login_required
def create_task():
<<<<<<< HEAD
    project_id = request.args.get('project_id')  # Проверяем, передан ли ID проекта в запросе
    project = None

    if project_id:
        project = Project.query.get_or_404(project_id)

        # Проверяем, что пользователь имеет доступ к проекту
        if current_user.id != project.owner_id and current_user not in project.members:
            flash('У вас нет доступа для добавления задач.', 'error')
            return redirect(url_for('projects'))

=======
>>>>>>> 4c80bc660c306d2a5b2908cf97d88b29abcddb6f
    if request.method == 'POST':
        # Получение данных из формы
        title = request.form['title']
        description = request.form['description']
        due_date = datetime.strptime(request.form['due_date'], '%Y-%m-%d')
        difficulty = float(request.form['difficulty'])
        priority = request.form['priority']
        status = request.form['status']
        assigned_to_id = request.form.get('assigned_to')

        # Создание новой задачи
<<<<<<< HEAD
        new_task = Task(
            title=title,
            description=description,
            due_date=due_date,
            difficulty=difficulty,
            priority=priority,
            status=status,
            user_id=current_user.id,
            assigned_to_id=assigned_to_id,
            project_id=project.id if project else None
        )

        db.session.add(new_task)

        # Работа с файлами
        files = request.files.getlist('task_files[]')
        for file in files:
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(file_path)
=======
        new_task = Task(title=title, description=description, due_date=due_date, difficulty=difficulty,
                        priority=priority, status=status, user_id=current_user.id, assigned_to_id=assigned_to_id)

        db.session.add(new_task)
        db.session.commit()

        # Работа с файлами
        files = request.files.getlist('task_files[]')  # Получаем список файлов
        for file in files:
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)  # Безопасное имя файла
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(file_path)  # Сохранение файла в папку uploads
>>>>>>> 4c80bc660c306d2a5b2908cf97d88b29abcddb6f

                # Сохранение информации о файле в базе данных
                new_file = File(filename=filename, task_id=new_task.id)
                db.session.add(new_file)

        # Подтверждение транзакции
        db.session.commit()

<<<<<<< HEAD
        # Перенаправление в зависимости от контекста
        if project:
            return redirect(url_for('project_detail', project_id=project.id))
        else:
            return redirect(url_for('dashboard'))

    # Получение списка пользователей для выбора исполнителя
    users = project.members if project else User.query.all()

    return render_template('create_task.html', users=users, project=project)
=======
        return redirect(url_for('dashboard'))

    # Получение списка пользователей для выбора исполнителя
    users = User.query.all()
    return render_template('create_task.html', users=users)
>>>>>>> 4c80bc660c306d2a5b2908cf97d88b29abcddb6f


# Удаление задачи
@app.route('/task/delete/<int:task_id>', methods=['POST'])
@login_required
def delete_task(task_id):
    task = Task.query.get_or_404(task_id)
    if task.user != current_user and current_user.role.role_name != 'Admin':

        return redirect(url_for('dashboard'))
    db.session.delete(task)
    db.session.commit()

    return redirect(url_for('dashboard'))


@app.route('/delete_file/<int:file_id>', methods=['POST'])
@login_required
def delete_file(file_id):
    file = File.query.get_or_404(file_id)

    # Путь к файлу
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)

    try:
        # Удаляем файл из файловой системы, если он существует
        if os.path.exists(file_path):
            os.remove(file_path)
            print(f"Файл {file.filename} успешно удалён.")
        else:
            print(f"Файл {file.filename} не найден в {file_path}. Возможно, он был удалён ранее.")

        # Удаляем запись о файле из базы данных
        db.session.delete(file)
        db.session.commit()

        return jsonify({"status": "success"})

    except Exception as e:
        # Логируем ошибку и возвращаем ответ с ошибкой
        print(f"Ошибка при удалении файла {file.filename}: {e}")
        return jsonify({"status": "error", "message": "Ошибка при удалении файла"}), 500


@app.route('/task/edit/<int:task_id>', methods=['GET', 'POST'])
@login_required
def edit_task(task_id):
    # Получаем задачу по ID или возвращаем 404, если не найдена
    task = Task.query.get_or_404(task_id)

    if request.method == 'POST':
        # Обновляем информацию о задаче
        task.title = request.form['title']
        task.description = request.form['description']
        task.due_date = datetime.strptime(request.form['due_date'], '%Y-%m-%d')
        task.difficulty = float(request.form['difficulty'])
        task.priority = request.form['priority']
        task.status = request.form['status']
        task.assigned_to_id = request.form.get('assigned_to')

        # Работа с файлами (новые загружаемые файлы)
        files = request.files.getlist('task_files[]')  # Проверяем правильность атрибута 'task_files[]'
        for file in files:
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(file_path)

                # Сохраняем информацию о новом файле в базе данных
                new_file = File(filename=filename, task_id=task.id)
                db.session.add(new_file)

        # Сохраняем изменения в базе данных
        db.session.commit()

        return redirect(url_for('dashboard'))

    # Получаем существующие файлы, привязанные к задаче
    files = File.query.filter_by(task_id=task_id).all()

    # Получаем список всех пользователей для выбора исполнителя
    users = User.query.all()

    # Передаем задачу, пользователей и файлы в шаблон для отображения
    return render_template('edit_task.html', task=task, users=users, files=files)

@app.route('/tasks', methods=['GET', 'POST'])
@login_required
def tasks():
    users = User.query.all()  # Load all users for the dropdown
    filters = {}

    if request.method == 'POST':
        if request.form.get('title'):
            filters['title'] = request.form['title']
        if request.form.get('assigned_to'):
            filters['assigned_to_id'] = int(request.form['assigned_to'])  # Get assigned user ID
        if request.form.get('priority'):
            filters['priority'] = request.form['priority']
        if request.form.get('due_date'):
            filters['due_date'] = datetime.strptime(request.form['due_date'], '%Y-%m-%d')
        if request.form.get('difficulty'):
            filters['difficulty'] = float(request.form['difficulty'])

        # Получение файлов из формы
        if 'task_files' in request.files:
            files = request.files.getlist('task_files')
            file_paths = []

            for file in files:
                if file and allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    filepath = os.path.join(UPLOAD_FOLDER, filename)
                    file.save(filepath)
                    file_paths.append(filepath)

        # Дополнительно, можно сохранить ссылки на файлы в базу данных или другой способ

        # Поиск задач
        tasks_query = Task.query

        if 'title' in filters:
            tasks_query = tasks_query.filter(Task.title.ilike(f"%{filters['title']}%"))
        if 'assigned_to_id' in filters:
            tasks_query = tasks_query.filter(Task.assigned_to_id == filters['assigned_to_id'])
        if 'priority' in filters:
            tasks_query = tasks_query.filter(Task.priority == filters['priority'])
        if 'due_date' in filters:
            tasks_query = tasks_query.filter(Task.due_date == filters['due_date'])
        if 'difficulty' in filters:
            tasks_query = tasks_query.filter(Task.difficulty == filters['difficulty'])

        tasks = tasks_query.all()

    else:
        tasks = Task.query.all()

    return render_template('tasks.html', tasks=tasks, users=users)


# Устанавливаем локаль для русского языка
locale.setlocale(locale.LC_TIME, 'Russian_Russia')
# Календарь задач
@app.route('/calendar', methods=['GET'])
@login_required
def calendar():
    current_year = request.args.get('year', datetime.now().year, type=int)
    current_month = request.args.get('month', datetime.now().month, type=int)

    if current_month > 12:
        current_month = 1
        current_year += 1
    elif current_month < 1:
        current_month = 12
        current_year -= 1

    if current_user.role.role_name == 'Admin':
        tasks = Task.query.filter(
            db.extract('year', Task.due_date) == current_year,
            db.extract('month', Task.due_date) == current_month
        ).all()
    else:
        tasks = Task.query.filter(
            (Task.user_id == current_user.id) | (Task.assigned_to_id == current_user.id),
            db.extract('year', Task.due_date) == current_year,
            db.extract('month', Task.due_date) == current_month
        ).all()

    days_in_month = monthcalendar(current_year, current_month)
    calendar_days = []

    for week in days_in_month:
        week_data = []
        for day in week:
            if day != 0:
                date = datetime(current_year, current_month, day)
                tasks_for_day = [task for task in tasks if task.due_date.date() == date.date()]
                week_data.append({'date': date, 'tasks': tasks_for_day})
            else:
                week_data.append({'date': None, 'tasks': []})
        calendar_days.append(week_data)

    current_month_name = month_name[current_month]

    previous_month = current_month - 1 if current_month > 1 else 12
    next_month = current_month + 1 if current_month < 12 else 1
    previous_year = current_year if current_month > 1 else current_year - 1
    next_year = current_year if current_month < 12 else current_year + 1

    # Подсчёт задач по сложности
    easy_count = Task.query.filter(
        Task.difficulty == 1,
        db.extract('year', Task.due_date) == current_year,
        db.extract('month', Task.due_date) == current_month
    ).count()

    medium_count = Task.query.filter(
        Task.difficulty == 2,
        db.extract('year', Task.due_date) == current_year,
        db.extract('month', Task.due_date) == current_month
    ).count()

    hard_count = Task.query.filter(
        Task.difficulty == 3,
        db.extract('year', Task.due_date) == current_year,
        db.extract('month', Task.due_date) == current_month
    ).count()

    # Подсчёт задач по приоритетам
    low_priority_count = sum(1 for task in tasks if task.priority == 'Низкий')
    medium_priority_count = sum(1 for task in tasks if task.priority == 'Средний')
    high_priority_count = sum(1 for task in tasks if task.priority == 'Высокий')

    return render_template('calendar.html',
                           calendar_days=calendar_days,
                           current_month_name=current_month_name,
                           current_year=current_year,
                           current_month=current_month,
                           previous_month=previous_month,
                           next_month=next_month,
                           previous_year=previous_year,
                           next_year=next_year,
                           easy_count=easy_count,
                           medium_count=medium_count,
                           hard_count=hard_count,
                           low_priority_count=low_priority_count,
                           medium_priority_count=medium_priority_count,
                           high_priority_count=high_priority_count)


@app.before_request
def check_deadlines():
    if current_user.is_authenticated:
        # Обновляем статус задач на "Просрочено", если срок истек
        tasks_in_progress = Task.query.filter_by(status='In Progress').all()
        for task in tasks_in_progress:
            if task.due_date < datetime.utcnow():
                task.status = 'Просрочено'
                db.session.commit()

        # Уведомление за 2 дня до дедлайна
        tasks = Task.query.filter_by(user_id=current_user.id, status='In Progress').all()
        for task in tasks:
            if task.due_date - datetime.utcnow() <= timedelta(days=2):
                flash(f'Задача "{task.title}" приближается к дедлайну!')


# Отчеты (KPI)
@app.route('/reports', methods=['GET', 'POST'])
@login_required
def reports():
    users = []
    # Добавляем список пользователей только для администратора
    if current_user.role.role_name == 'Admin':
        users = User.query.all()

    # Выбор задач для вычисления KPI
    if request.method == 'POST':
        if current_user.role.role_name == 'Admin':
            # Админ может выбрать другого пользователя
            user_id = request.form.get('user_id')
            tasks = Task.query.filter(
                (Task.user_id == user_id) | (Task.assigned_to_id == user_id)
            ).all()
        else:
            # Для обычных пользователей выбираем только их задачи
            tasks = Task.query.filter(
                (Task.user_id == current_user.id) | (Task.assigned_to_id == current_user.id)
            ).all()
    else:
        # Если не POST запрос, то показываем только задачи текущего пользователя
        tasks = Task.query.filter(
            (Task.user_id == current_user.id) | (Task.assigned_to_id == current_user.id)
        ).all()

    # Личный KPI для текущего пользователя (или выбранного администратором)
    user_kpi = calculate_kpi(tasks)

    # Общий KPI для всех задач в системе (не зависит от пользователя)
    all_tasks = Task.query.all()
    company_kpi = calculate_kpi(all_tasks)

    return render_template('reports.html', user_kpi=user_kpi, company_kpi=company_kpi, users=users)


def calculate_kpi(tasks):
<<<<<<< HEAD
    """
    Пересчитанная функция KPI с учётом:
    - сложности
    - приоритета
    - своевременности выполнения
    """
    difficulty_mapping = {'Легко': 1, 'Средне': 2, 'Сложно': 3}
    priority_mapping = {'Низкий': 0.5, 'Средний': 1, 'Высокий': 1.5}
    overdue_penalty = 0.5  # Штраф за просрочку
    late_completion_penalty = 0.8  # Штраф за выполнение с опозданием (80% от веса задачи)

    total_weighted_value = 0  # Общий вес задач
    completed_on_time_value = 0  # Вес задач, завершённых вовремя
    completed_late_value = 0  # Вес задач, завершённых с опозданием
    overdue_value = 0  # Вес просроченных задач

    now = datetime.utcnow()

    for task in tasks:
        # Определяем вес задачи по сложности и приоритету
=======
    """Функция для расчета KPI на основе списка задач"""
    difficulty_mapping = {'Легко': 1, 'Средне': 2, 'Сложно': 3}
    priority_mapping = {'Низкий': 0.5, 'Средний': 1, 'Высокий': 1.5}
    overdue_penalty = 0.5

    total_weighted_value = 0
    completed_weighted_value = 0
    overdue_weighted_value = 0

    for task in tasks:
>>>>>>> 4c80bc660c306d2a5b2908cf97d88b29abcddb6f
        task_difficulty = difficulty_mapping.get(task.difficulty, 1)
        task_priority = priority_mapping.get(task.priority, 1)
        weighted_value = task_difficulty * task_priority

        total_weighted_value += weighted_value

<<<<<<< HEAD
        # Определяем KPI в зависимости от статуса задачи
        if task.status == 'Completed':
            if task.due_date >= now:  # Завершена вовремя
                completed_on_time_value += weighted_value
            else:  # Завершена с опозданием
                completed_late_value += weighted_value * late_completion_penalty
        elif task.status == 'Просрочено':  # Ещё не завершена, но уже просрочена
            overdue_value += weighted_value * overdue_penalty

    # Считаем итоговый KPI
    if total_weighted_value > 0:
        kpi_score = ((completed_on_time_value + completed_late_value) / total_weighted_value) * 100
    else:
        kpi_score = 0

    return round(kpi_score, 2)

=======
        if task.status == 'Completed':
            completed_weighted_value += weighted_value
        elif task.status == 'Просрочено':
            overdue_weighted_value += weighted_value * overdue_penalty

    # Вместо того, чтобы добавлять штраф, используем это отдельно
    completed_weighted_value += overdue_weighted_value

    # Измените логику KPI, чтобы избежать деления на 0
    if total_weighted_value > 0:
        kpi_score = (completed_weighted_value / total_weighted_value) * 100
    else:
        kpi_score = 0

    return kpi_score
>>>>>>> 4c80bc660c306d2a5b2908cf97d88b29abcddb6f


@app.route('/download_report', methods=['POST'])
@login_required
def download_report():
    start_date_str = request.form.get('start_date')
    end_date_str = request.form.get('end_date')
    selected_user_id = request.form.get('user_id')

    start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d')

    # Если админ, выбирает пользователя, иначе — текущий пользователь
    if current_user.role.role_name == 'Admin' and selected_user_id:
        user_id = int(selected_user_id)
    else:
        user_id = current_user.id

    tasks = Task.query.filter(
        (Task.user_id == user_id) | (Task.assigned_to_id == user_id)
    ).filter(Task.due_date.between(start_date, end_date)).all()

    difficulty_mapping = {'Легко': 1, 'Средне': 2, 'Сложно': 3}
    priority_mapping = {'Низкий': 0.5, 'Средний': 1, 'Высокий': 1.5}
    overdue_penalty = 0.5
<<<<<<< HEAD
    late_completion_penalty = 0.8

    total_weighted_value = 0
    completed_on_time_value = 0
    completed_late_value = 0
    overdue_value = 0

    for task in tasks:
        # Определяем вес задачи по сложности и приоритету
=======

    total_weighted_value = 0
    completed_weighted_value = 0

    for task in tasks:
>>>>>>> 4c80bc660c306d2a5b2908cf97d88b29abcddb6f
        task_difficulty = difficulty_mapping.get(task.difficulty, 1)
        task_priority = priority_mapping.get(task.priority, 1)
        weighted_value = task_difficulty * task_priority

        total_weighted_value += weighted_value

        if task.status == 'Completed':
<<<<<<< HEAD
            if task.due_date >= datetime.utcnow():
                completed_on_time_value += weighted_value
            else:
                completed_late_value += weighted_value * late_completion_penalty
        elif task.status == 'Просрочено':
            overdue_value += weighted_value * overdue_penalty

    if total_weighted_value > 0:
        kpi_score = ((completed_on_time_value + completed_late_value) / total_weighted_value) * 100
=======
            completed_weighted_value += weighted_value
        elif task.status == 'Просрочено':
            completed_weighted_value += weighted_value * overdue_penalty

    if total_weighted_value > 0:
        kpi_score = (completed_weighted_value / total_weighted_value) * 100
>>>>>>> 4c80bc660c306d2a5b2908cf97d88b29abcddb6f
    else:
        kpi_score = 0

    # Создаем Excel файл
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Отчет"

    headers = ['Название', 'Описание', 'Приоритет', 'Статус', 'Крайняя дата выполнения', 'Сложность']
    sheet.append(headers)

    for task in tasks:
        sheet.append([task.title, task.description, task.priority, task.status, task.due_date.strftime('%Y-%m-%d'), task.difficulty])

    sheet.append([])
    sheet.append(['KPI за выбранный период:', round(kpi_score, 2)])

    for col_num, col_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        sheet.column_dimensions[column_letter].width = 20

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'report_{start_date_str}_to_{end_date_str}.xlsx'
    )
from flask import send_file
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
import io

@app.route('/download_all_kpis', methods=['POST'])
@login_required
def download_all_kpis():
    # Получение периода
    start_date_str = request.form.get('start_date')
    end_date_str = request.form.get('end_date')
    start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d')

    # Проверка роли
    if current_user.role.role_name != 'Admin':
        flash("У вас нет прав для скачивания общего отчета.")
        return redirect(url_for('reports'))

    # Сбор KPI всех сотрудников
    users = User.query.all()
    user_kpis = []
    for user in users:
        tasks = Task.query.filter(
            (Task.user_id == user.id) | (Task.assigned_to_id == user.id)
        ).filter(Task.due_date.between(start_date, end_date)).all()

        # Расчет KPI
        kpi_score = calculate_kpi(tasks)
        user_kpis.append({'user': user.username, 'kpi': round(kpi_score, 2)})

    # Создание Excel-файла
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Общий KPI отчет"

    headers = ['Сотрудник', 'KPI']
    sheet.append(headers)

    for user_kpi in user_kpis:
        sheet.append([user_kpi['user'], user_kpi['kpi']])

    # Добавление диаграммы
    chart = BarChart()
    chart.title = "KPI сотрудников"
    chart.x_axis.title = "Сотрудники"
    chart.y_axis.title = "KPI (%)"

    data = Reference(sheet, min_col=2, min_row=2, max_row=len(user_kpis) + 1)
    categories = Reference(sheet, min_col=1, min_row=2, max_row=len(user_kpis) + 1)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(categories)
    sheet.add_chart(chart, "D4")

    # Форматирование столбцов
    for col_num, col_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        sheet.column_dimensions[column_letter].width = 20

    # Сохранение файла
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'all_kpis_{start_date_str}_to_{end_date_str}.xlsx'
    )


# Маршруты для диаграммы Ганта
from datetime import datetime
from flask import Flask, jsonify, request

# Маршрут для диаграммы Ганта
@app.route('/gantt', methods=['GET', 'POST'])
@login_required
def gantt():
    if request.method == 'POST':
        title = request.form['title']
        description = request.form['description']
        due_date = datetime.strptime(request.form['due_date'], '%Y-%m-%d')
        priority = request.form['priority']
        status = request.form['status']
        difficulty = request.form['difficulty']
        assigned_to_id = request.form.get('assigned_to')

        new_task = Task(
            title=title,
            description=description,
            due_date=due_date,
            priority=priority,
            status=status,
            difficulty=difficulty,
            user_id=current_user.id,
            assigned_to_id=assigned_to_id
        )
        db.session.add(new_task)
        db.session.commit()
        return redirect(url_for('gantt'))

    tasks = Task.query.all()  # Загружаем все задачи
    tasks_data = [task.to_dict() for task in tasks]  # Преобразуем задачи в формат JSON
    users = User.query.all()  # Пользователи для выбора ответственного
    return render_template('gantt.html', tasks_data=tasks_data, users=users)

# Добавление подзадачи
# Добавление подзадачи
@app.route('/add_subtask/<int:task_id>', methods=['POST'])
@login_required
def add_subtask(task_id):
    try:
        data = request.get_json()
        start_date = datetime.strptime(data['start_date'], '%Y-%m-%d')
        end_date = datetime.strptime(data['end_date'], '%Y-%m-%d')

        new_subtask = SubTask(
            task_id=task_id,
            title=data['title'],
            start_date=start_date,
            end_date=end_date
        )
        db.session.add(new_subtask)
        db.session.commit()
        return jsonify({'status': 'success', 'subtask': new_subtask.to_dict()})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 400

# Удаление подзадачи
@app.route('/delete_subtask/<int:subtask_id>', methods=['DELETE'])
@login_required
def delete_subtask(subtask_id):
    subtask = SubTask.query.get(subtask_id)
    if subtask:
        db.session.delete(subtask)
        db.session.commit()
        return jsonify({'status': 'success'})
    return jsonify({'status': 'error', 'message': 'Подзадача не найдена'}), 404

# Создание новой подзадачи
@app.route('/create_subtask', methods=['POST'])
@login_required
def create_subtask():
    try:
        data = request.get_json()
        start_date = datetime.strptime(data['start_date'], '%Y-%m-%d')
        end_date = datetime.strptime(data['end_date'], '%Y-%m-%d')

        new_subtask = SubTask(
            task_id=data['task_id'],
            title=data['title'],
            start_date=start_date,
            end_date=end_date
        )
        db.session.add(new_subtask)
        db.session.commit()
        return jsonify({'status': 'success', 'subtask': new_subtask.to_dict()})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 400

# Обновление подзадачи
@app.route('/update_subtask', methods=['POST'])
@login_required
def update_subtask():
    try:
        data = request.get_json()
        subtask = SubTask.query.get(data['id'])
        if not subtask:
            return jsonify({'status': 'error', 'message': 'Подзадача не найдена'}), 404

        subtask.title = data['title']
        subtask.start_date = datetime.strptime(data['start_date'], '%Y-%m-%d')
        subtask.end_date = datetime.strptime(data['end_date'], '%Y-%m-%d')
        db.session.commit()
        return jsonify({'status': 'success', 'subtask': subtask.to_dict()})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 400

# Добавление комментария
# Добавление комментария
@app.route('/task/<int:task_id>/add_comment', methods=['POST'])
@login_required
def add_comment(task_id):
    try:
        # Получаем данные из запроса
        data = request.get_json()
        comment_text = data.get('comment', '').strip()
        if not comment_text:
            return jsonify({'status': 'error', 'message': 'Комментарий не может быть пустым'}), 400

        # Проверяем существование задачи
        task = Task.query.get(task_id)
        if not task:
            return jsonify({'status': 'error', 'message': 'Задача не найдена'}), 404

        # Создаем комментарий
        new_comment = TaskComments(
            task_id=task_id,
            user_id=current_user.id,
            comment=comment_text
        )
        db.session.add(new_comment)
        db.session.commit()

        # Возвращаем добавленный комментарий
        return jsonify({
            'status': 'success',
            'comment': new_comment.to_dict()
        }), 200
    except Exception as e:
        app.logger.error(f"Ошибка при добавлении комментария: {e}")
        return jsonify({'status': 'error', 'message': 'Внутренняя ошибка сервера'}), 500

# Редактирование комментария
@app.route('/task/<int:task_id>/edit_comment/<int:comment_id>', methods=['PUT'])
@login_required
def edit_comment(task_id, comment_id):
    try:
        data = request.get_json()
        new_text = data.get('comment', '').strip()
        if not new_text:
            return jsonify({'status': 'error', 'message': 'Комментарий не может быть пустым'}), 400

        comment = TaskComments.query.filter_by(id=comment_id, task_id=task_id).first()
        if not comment:
            return jsonify({'status': 'error', 'message': 'Комментарий не найден'}), 404
        if comment.user_id != current_user.id:
            return jsonify({'status': 'error', 'message': 'У вас нет прав на редактирование этого комментария'}), 403

        comment.comment = new_text
        db.session.commit()

        return jsonify({
            'status': 'success',
            'comment': {
                'id': comment.id,
                'task_id': comment.task_id,
                'user_id': comment.user_id,
                'comment': comment.comment,
                'created_at': comment.created_at.strftime('%Y-%m-%d %H:%M:%S')
            }
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 400


# Удаление комментария
@app.route('/task/<int:task_id>/delete_comment/<int:comment_id>', methods=['DELETE'])
@login_required
def delete_comment(task_id, comment_id):
    try:
        comment = TaskComments.query.filter_by(id=comment_id, task_id=task_id).first()
        if not comment:
            return jsonify({'status': 'error', 'message': 'Комментарий не найден'}), 404
        if comment.user_id != current_user.id:
            return jsonify({'status': 'error', 'message': 'У вас нет прав на удаление этого комментария'}), 403

        db.session.delete(comment)
        db.session.commit()

        return jsonify({'status': 'success', 'message': 'Комментарий удалён'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 400

    from flask import jsonify

    @app.route('/task/<int:task_id>')
    @login_required
    def get_task(task_id):
        task = Task.query.get_or_404(task_id)
        comments = [
            {
                'id': c.id,
                'comment': c.comment,
                'user_id': c.user_id,
                'created_at': c.created_at.strftime('%Y-%m-%d %H:%M:%S')
            }
            for c in task.task_comments
        ]
        return jsonify({'id': task.id, 'title': task.title, 'comments': comments})
# Конфигурация Document Server
DOCUMENT_SERVER = "http://localhost"  # URL OnlyOffice Document Server

# Папка для загрузки файлов
UPLOAD_FOLDER = os.path.join(os.getcwd(), 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
import os
from flask import Flask, render_template
from werkzeug.utils import secure_filename

UPLOAD_FOLDER = os.path.join(os.getcwd(), 'uploads')


@app.route('/editor', methods=['GET'])
def editor_page():
    # Получаем список файлов из папки uploads
    files = []
    for filename in os.listdir(UPLOAD_FOLDER):
        # Добавляем только файлы (игнорируем папки)
        if os.path.isfile(os.path.join(UPLOAD_FOLDER, filename)):
            files.append(filename)

    return render_template('editor.html', files=files)


from flask_login import current_user

UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'docx', 'xlsx', 'pptx', 'txt'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route('/edit/<filename>', methods=['GET'])
def edit_document(filename):
    # Уникальный ключ для документа
    document_key = md5(f"{filename}_{current_user.id}_{datetime.utcnow()}".encode()).hexdigest()

    file_url = f"http://127.0.0.1:5000/uploads/{filename}"
    callback_url = f"http://127.0.0.1:5000/callback"

    # Конфигурация документа
    document_config = {
        "document": {
            "fileType": filename.split('.')[-1].lower(),
            "key": document_key,
            "title": filename,
            "url": file_url,
            "permissions": {"edit": True}
        },
        "editorConfig": {
            "callbackUrl": callback_url,
            "mode": "edit",
            "lang": "ru",
            "user": {
                "id": current_user.id,
                "name": current_user.username
            }
        }
    }

    # Логирование конфигурации
    app.logger.info(f"Конфигурация документа: {document_config}")

    # Передача конфигурации в шаблон
    return render_template('editor_frame.html', document_config=document_config)


@app.route('/upload/<task_id>', methods=['POST'])
def upload_file_for_task(task_id):
    if 'file' not in request.files:
        return "No file part", 400
    file = request.files['file']
    if file.filename == '':
        return "No selected file", 400
    if not allowed_file(file.filename):
        return "Недопустимый формат файла.", 400

    file_path = os.path.join(UPLOAD_FOLDER, file.filename)
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)  # Создаём папку, если её нет
    file.save(file_path)

    # Логирование загрузки файла
    app.logger.info(f"Файл {file.filename} успешно загружен в {UPLOAD_FOLDER}")

    # Добавляем файл в базу данных
    new_file = File(filename=file.filename, task_id=task_id)
    db.session.add(new_file)
    db.session.commit()

    return redirect(url_for('editor_page', task_id=task_id))


@app.route('/uploads/<filename>')
def uploaded_file(filename):
    file_path = os.path.join(UPLOAD_FOLDER, filename)

    # Проверка существования файла
    if not os.path.exists(file_path):
        app.logger.error(f"Файл {filename} не найден в {UPLOAD_FOLDER}.")
        return "Файл не найден", 404

    # Определение MIME-типа
    mime_type, _ = mimetypes.guess_type(filename)
    if not mime_type:
        mime_type = 'application/octet-stream'

    # Логирование отправки файла
    app.logger.info(f"Отправка файла {filename} с MIME-типом {mime_type}")
    return send_from_directory(UPLOAD_FOLDER, filename, mimetype=mime_type)


@app.route('/callback', methods=['POST'])
def callback():
    data = request.json
    app.logger.info(f"Callback received: {data}")
    return jsonify({"error": 0})  # Всегда возвращайте успешный ответ

<<<<<<< HEAD
@app.route('/reminders', methods=['GET'])
@login_required
def view_reminders():
    reminders = Reminder.query.filter_by(user_id=current_user.id).order_by(Reminder.reminder_date).all()
    return render_template('reminders.html', reminders=reminders)

@app.route('/reminder/new', methods=['GET', 'POST'])
@login_required
def create_reminder():
    if request.method == 'POST':
        title = request.form['title']
        description = request.form.get('description', '')
        # Изменённый формат для datetime-local
        reminder_date = datetime.strptime(request.form['reminder_date'], '%Y-%m-%dT%H:%M')
        priority = request.form['priority']
        repeat = request.form['repeat']

        new_reminder = Reminder(
            title=title,
            description=description,
            reminder_date=reminder_date,
            priority=priority,
            repeat=repeat,
            user_id=current_user.id
        )
        db.session.add(new_reminder)
        db.session.commit()
        flash('Напоминание успешно создано!', 'success')
        return redirect(url_for('view_reminders'))

    return render_template('create_reminder.html')


@app.route('/reminder/edit/<int:reminder_id>', methods=['GET', 'POST'])
@login_required
def edit_reminder(reminder_id):
    reminder = Reminder.query.filter_by(id=reminder_id, user_id=current_user.id).first_or_404()

    if request.method == 'POST':
        reminder.title = request.form['title']
        reminder.description = request.form.get('description', '')
        # Изменённый формат для datetime-local
        reminder.reminder_date = datetime.strptime(request.form['reminder_date'], '%Y-%m-%dT%H:%M')
        reminder.priority = request.form['priority']
        reminder.repeat = request.form['repeat']
        db.session.commit()
        flash('Напоминание успешно обновлено!', 'success')
        return redirect(url_for('view_reminders'))

    return render_template('edit_reminder.html', reminder=reminder)

@app.route('/reminder/delete/<int:reminder_id>', methods=['POST'])
@login_required
def delete_reminder(reminder_id):
    reminder = Reminder.query.filter_by(id=reminder_id, user_id=current_user.id).first_or_404()
    db.session.delete(reminder)
    db.session.commit()
    flash('Напоминание успешно удалено!', 'success')
    return redirect(url_for('view_reminders'))

@app.route('/task_board', methods=['GET'])
@login_required
def task_board():
    tasks = Task.query.filter_by(user_id=current_user.id).all()

    # Категоризация задач по выполненности
    tasks_by_category = {
        'В работе': [task for task in tasks if task.status == 'In Progress' and task.due_date >= datetime.utcnow()],
        'Выполненные': [task for task in tasks if task.status == 'Completed'],
        'Просроченные': [task for task in tasks if task.status == 'Просрочено'],  # Ищем статус 'Просрочено'
    }

    return render_template(
        'task_board.html',
        tasks_by_category=tasks_by_category,
        board_type="Готовность"
    )

@app.before_request
def check_deadlines():
    if current_user.is_authenticated:
        # Обновляем статус задач на "Просрочено", если срок истек
        tasks_in_progress = Task.query.filter_by(status='In Progress').all()
        for task in tasks_in_progress:
            if task.due_date < datetime.utcnow():
                task.status = 'Просрочено'  # Меняем статус
        db.session.commit()


@app.route('/task_board/priority', methods=['GET'])
@login_required
def task_board_priority():
    tasks = Task.query.filter_by(user_id=current_user.id).all()

    # Маппинг приоритетов
    tasks_by_priority = {
        'Низкий': [task for task in tasks if task.priority == 'Низкий'],
        'Средний': [task for task in tasks if task.priority == 'Средний'],
        'Высокий': [task for task in tasks if task.priority == 'Высокий'],
    }

    return render_template('task_board.html', tasks_by_category=tasks_by_priority, board_type="Приоритет")


@app.route('/task_board/difficulty', methods=['GET'])
@login_required
def task_board_difficulty():
    tasks = Task.query.filter_by(user_id=current_user.id).all()

    # Маппинг сложности
    difficulty_mapping = {1: 'Меньше 2 часов', 2: 'От 2 до 8 часов', 3: 'Более 8 часов'}
    tasks_by_difficulty = {
        'Меньше 2 часов': [task for task in tasks if task.difficulty == 1],
        'От 2 до 8 часов': [task for task in tasks if task.difficulty == 2],
        'Более 8 часов': [task for task in tasks if task.difficulty == 3],
    }

    return render_template(
        'task_board.html',
        tasks_by_category=tasks_by_difficulty,
        board_type="Сложность"
    )


@app.route('/update_task_category', methods=['POST'])
@login_required
def update_task_category():
    data = request.json
    task_id = data.get('task_id')
    new_category = data.get('new_category')

    # Получаем задачу
    task = Task.query.filter_by(id=task_id, user_id=current_user.id).first()
    if not task:
        return jsonify({'error': 'Задача не найдена'}), 404

    # Определяем, что обновлять
    if request.referrer.endswith('/task_board'):  # Доска выполненности
        if new_category == 'В работе':
            task.status = 'In Progress'
        elif new_category == 'Выполненные':
            task.status = 'Completed'
        elif new_category == 'Просроченные':
            task.status = 'Просрочено'

    elif request.referrer.endswith('/task_board/priority'):  # Доска приоритетов

        task.priority = new_category

    elif request.referrer.endswith('/task_board/difficulty'):  # Доска сложности
        difficulty_mapping = {'Меньше 2 часов': 1, 'От 2 до 8 часов': 2, 'Более 8 часов': 3}
        task.difficulty = difficulty_mapping.get(new_category)

    # Сохраняем изменения
    db.session.commit()
    return jsonify({'success': True})
@app.route('/projects', methods=['GET'])
@login_required
def projects():
    projects = Project.query.filter(
        (Project.owner_id == current_user.id) |
        (Project.members.any(id=current_user.id))
    ).all()
    return render_template('projects.html', projects=projects)

@app.route('/project/new', methods=['GET', 'POST'])
@login_required
def create_project():
    if request.method == 'POST':
        name = request.form['name']
        description = request.form.get('description', '')
        new_project = Project(name=name, description=description, owner_id=current_user.id)
        db.session.add(new_project)
        db.session.commit()
        flash('Проект успешно создан!', 'success')
        return redirect(url_for('projects'))
    return render_template('create_project.html')

@app.route('/project/<int:project_id>', methods=['GET'])
def project_detail(project_id):
    project = Project.query.get_or_404(project_id)
    all_users = User.query.all()  # Получаем список всех пользователей
    return render_template('project_detail.html', project=project, all_users=all_users)
    tasks = Task.query.filter_by(project_id=project.id).all()
    return render_template('project_detail.html', project=project, tasks=tasks)

@app.route('/project/<int:project_id>/add_member', methods=['POST'])
def add_project_member(project_id):
    user_id = request.form['user_id']
    project = Project.query.get_or_404(project_id)
    user = User.query.get(user_id)  # Находим пользователя по ID
    if user and project.owner_id == current_user.id:  # Проверка прав владельца
        project.members.append(user)  # Добавляем пользователя в проект
        db.session.commit()  # Сохраняем изменения
        flash('Участник добавлен в проект!', 'success')
    return redirect(url_for('project_detail', project_id=project_id))

@app.route('/project/<int:project_id>/task/new', methods=['GET', 'POST'])
@login_required
def create_task_in_project(project_id):
    project = Project.query.get_or_404(project_id)
    if current_user.id != project.owner_id and current_user not in project.members:
        flash('У вас нет доступа для добавления задач.', 'error')
        return redirect(url_for('project_detail', project_id=project_id))

    if request.method == 'POST':
        title = request.form['title']
        description = request.form['description']
        priority = request.form['priority']
        status = request.form['status']
        difficulty = float(request.form['difficulty'])
        due_date = datetime.strptime(request.form['due_date'], '%Y-%m-%d')
        assigned_to_id = request.form.get('assigned_to')

        new_task = Task(
            title=title,
            description=description,
            priority=priority,
            status=status,
            difficulty=difficulty,
            due_date=due_date,
            project_id=project_id,
            user_id=current_user.id,
            assigned_to_id=assigned_to_id
        )
        db.session.add(new_task)
        db.session.commit()
        flash('Задача успешно добавлена!', 'success')
        return redirect(url_for('project_detail', project_id=project_id))

    # Передаём только участников проекта
    return render_template('create_task.html', project=project, users=project.members)

def project_access_required(func):
    @wraps(func)
    def decorated_view(*args, **kwargs):
        project_id = kwargs.get('project_id')
        project = Project.query.get_or_404(project_id)
        if current_user.id != project.owner_id and current_user not in project.members:
            flash('У вас нет доступа к этому проекту.', 'error')
            return redirect(url_for('projects'))
        return func(*args, **kwargs)
    return decorated_view

@app.route('/project/<int:project_id>/delete', methods=['POST'])
@login_required
def delete_project(project_id):
    project = Project.query.get_or_404(project_id)
    if project.owner_id != current_user.id:
        flash('Только владелец проекта может удалить его.', 'error')
        return redirect(url_for('project_detail', project_id=project_id))
    db.session.delete(project)
    db.session.commit()
    flash('Проект успешно удалён.', 'success')
    return redirect(url_for('projects'))
@app.route('/project/<int:project_id>/remove_member', methods=['POST'])
@login_required
def remove_project_member(project_id):
    project = Project.query.get_or_404(project_id)
    if project.owner_id != current_user.id:
        flash('Только владелец проекта может удалять участников.', 'error')
        return redirect(url_for('project_detail', project_id=project_id))
    user_id = request.form.get('user_id')
    user = User.query.get(user_id)
    if user and user in project.members:
        project.members.remove(user)
        db.session.commit()
        flash('Участник удалён.', 'success')
    return redirect(url_for('project_detail', project_id=project_id))
@app.route('/project/<int:project_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_project(project_id):
    project = Project.query.get_or_404(project_id)
    if project.owner_id != current_user.id:
        flash('Только владелец проекта может редактировать его.', 'error')
        return redirect(url_for('project_detail', project_id=project_id))

    if request.method == 'POST':
        project.name = request.form['name']
        project.description = request.form.get('description', '')
        db.session.commit()
        flash('Проект успешно обновлён.', 'success')
        return redirect(url_for('project_detail', project_id=project_id))
    return render_template('edit_project.html', project=project)



@app.before_request
def check_deadlines():
    if current_user.is_authenticated:
        # Обновляем статус задач на "Просрочено", если срок истек
        tasks_in_progress = Task.query.filter_by(status='In Progress').all()
        for task in tasks_in_progress:
            if task.due_date < datetime.utcnow():
                task.status = 'Просрочено'  # Меняем статус на "Просрочено"
        db.session.commit()




@app.route('/project_tasks/<int:project_id>', methods=['GET'])
@login_required
def project_tasks(project_id):
    # Получаем объект проекта по project_id
    project = Project.query.get_or_404(project_id)

    # Получаем задачи проекта по project_id
    tasks = Task.query.filter_by(project_id=project_id).all()

    # Категоризация задач по готовности
    project_tasks_by_category = {
        'В работе': [task for task in tasks if task.status == 'In Progress' and task.due_date >= datetime.utcnow()],
        'Выполненные': [task for task in tasks if task.status == 'Completed'],
        'Просроченные': [task for task in tasks if task.status == 'Просрочено'],
    }

    return render_template(
        'project_tasks.html',
        project=project,  # Передаем объект проекта в шаблон
        project_tasks_by_category=project_tasks_by_category,
        board_type="Готовность",
        project_id=project_id
    )

@app.route('/project_tasks_priority/<int:project_id>', methods=['GET'])
@login_required
def project_tasks_priority(project_id):
    # Получаем объект проекта по project_id
    project = Project.query.get_or_404(project_id)

    # Получаем задачи проекта по project_id
    tasks = Task.query.filter_by(project_id=project_id).all()

    # Категоризация задач по приоритету
    project_tasks_by_priority = {
        'Низкий': [task for task in tasks if task.priority == 'Низкий'],
        'Средний': [task for task in tasks if task.priority == 'Средний'],
        'Высокий': [task for task in tasks if task.priority == 'Высокий'],
    }

    return render_template(
        'project_tasks.html',
        project=project,  # Передаем объект проекта в шаблон
        project_tasks_by_category=project_tasks_by_priority,
        board_type="Приоритет",
        project_id=project_id
    )

@app.route('/project_tasks_difficulty/<int:project_id>', methods=['GET'])
@login_required
def project_tasks_difficulty(project_id):
    # Получаем объект проекта по project_id
    project = Project.query.get_or_404(project_id)

    # Получаем задачи проекта по project_id
    tasks = Task.query.filter_by(project_id=project_id).all()

    difficulty_mapping = {1: 'Меньше 2 часов', 2: 'От 2 до 8 часов', 3: 'Более 8 часов'}
    project_tasks_by_difficulty = {
        'Меньше 2 часов': [task for task in tasks if task.difficulty == 1],
        'От 2 до 8 часов': [task for task in tasks if task.difficulty == 2],
        'Более 8 часов': [task for task in tasks if task.difficulty == 3],
    }

    return render_template(
        'project_tasks.html',
        project=project,  # Передаем объект проекта в шаблон
        project_tasks_by_category=project_tasks_by_difficulty,
        board_type="Сложность",
        project_id=project_id
    )

@app.route('/update_project_task_category', methods=['POST'])
@login_required
def update_project_task_category():
    data = request.json
    task_id = data.get('task_id')
    new_category = data.get('new_category')

    # Получаем задачу
    task = Task.query.filter_by(id=task_id, user_id=current_user.id).first()
    if not task:
        return jsonify({'error': 'Задача не найдена'}), 404

    # Обновление статуса, приоритета или сложности в зависимости от категории
    if new_category in ['В работе', 'Выполненные', 'Просроченные']:
        if new_category == 'В работе':
            task.status = 'In Progress'
        elif new_category == 'Выполненные':
            task.status = 'Completed'
        elif new_category == 'Просроченные':
            task.status = 'Overdue'

    elif new_category in ['Низкий', 'Средний', 'Высокий']:
        task.priority = new_category

    elif new_category in ['Меньше 2 часов', 'От 2 до 8 часов', 'Более 8 часов']:
        difficulty_mapping = {'Меньше 2 часов': 1, 'От 2 до 8 часов': 2, 'Более 8 часов': 3}
        task.difficulty = difficulty_mapping.get(new_category)

    # Сохраняем изменения
    db.session.commit()
    return jsonify({'success': True})

import random
from flask import jsonify, render_template, request
from flask_login import login_required, current_user

@app.route('/roulette', methods=['GET', 'POST'])
@login_required
def roulette():
    # Рассчитываем количество доступных жетонов на основе выполненных задач
    completed_tasks = Task.query.filter_by(assigned_to_id=current_user.id, status='Completed').count()
    available_tokens = completed_tasks * 10

    # Определяем диапазоны
    ranges = [(i, i + 14) for i in range(0, 75, 15)]

    if request.method == 'POST':
        # Получаем ставку и тип ставки от клиента
        data = request.json
        bet_amount = int(data.get('bet_amount', 0))
        bet_type = data.get('bet_type', None)
        bet_value = data.get('bet_value', None)  # Например, цвет или число

        # Проверяем корректность ставки
        if bet_amount <= 0 or bet_amount > available_tokens:
            return jsonify({'status': 'error', 'message': 'Недостаточно жетонов или неверная ставка!'}), 400

        # Проверка на вероятность победы (3 из 20)
        win_probability = random.randint(1, 20)
        is_win = win_probability <= 3  # 3 победных результата из 20

        # Генерируем результат рулетки
        result = random.randint(0, 74)
        result_color = "зеленое" if result == 0 else ("красное" if result % 2 == 0 else "черное")

        # Проверяем, выиграл ли пользователь
        win_multiplier = 0

        if is_win:  # Если это выигрыш
            if bet_type == 'color':  # Ставка на цвет
                if bet_value == result_color:
                    win_multiplier = 2

            elif bet_type == 'range':  # Ставка на диапазон
                # Преобразуем ставку в диапазон
                try:
                    range_start, range_end = map(int, bet_value.split('-'))
                    if (range_start, range_end) in ranges and range_start <= result <= range_end:
                        win_multiplier = 5
                except ValueError:
                    return jsonify({'status': 'error', 'message': 'Неверный формат диапазона!'}), 400

            elif bet_type == 'number':  # Ставка на конкретное число
                if int(bet_value) == result:
                    win_multiplier = 75

        # Вычисляем выигрыш
        winnings = bet_amount * win_multiplier

        # Формируем ответ
        return jsonify({
            'status': 'success',
            'result': {
                'rolled_number': result,
                'rolled_color': result_color,
                'winnings': winnings,
                'bet_amount': bet_amount,
                'available_tokens': available_tokens - bet_amount + winnings
            },
            'message': f"Вы {'выиграли' if winnings > 0 else 'проиграли'}!"
        })

    # Возвращаем HTML и баланс жетонов при GET-запросе
    return render_template('roulette.html', tokens=available_tokens, ranges=ranges)
=======

>>>>>>> 4c80bc660c306d2a5b2908cf97d88b29abcddb6f

# Запуск приложения
if __name__ == '__main__':
    app.run(debug=True)
